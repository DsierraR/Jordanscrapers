import requests
from bs4 import BeautifulSoup
import re
import pandas as pd
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.image import MIMEImage
from openpyxl import load_workbook, Workbook
from datetime import datetime
import matplotlib.pyplot as plt
import pytz
import traceback
import logging
import boto3
from io import BytesIO
import numpy as np
from email.mime.base import MIMEBase
from email import encoders
from matplotlib.ticker import ScalarFormatter
from urllib.parse import urljoin

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# AWS Setup
s3_client = boto3.client(
    's3',
    aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
    aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
    region_name=os.getenv('AWS_REGION')
)
bucket_name = 'ctabucketdata'
file_key = '4scrapesdata.xlsx'

# Timezone for date on Excel
local_tz = pytz.timezone('America/Chicago')

# ETF Abbreviations
ETF_ABBREVIATIONS = {
    'American Beacon': 'AHL',
    'IMGP Funds': 'DBMF',
    'KFA Funds': 'KMLM',
    'Simplify': 'SY'
}

from urllib.parse import urljoin

def scrape_american_beacon():
    logging.info('Starting the scraping process for American Beacon...')
    url = 'https://www.americanbeaconfunds.com/etfs/ahlt.aspx'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    link_tag = soup.find('a', href=re.compile(r'\d{8}_AHLT_Holdings\.csv'))
    if not link_tag:
        logging.error("Could not find the CSV download link on the American Beacon page.")
        return pd.DataFrame()
    
    csv_link = link_tag['href']
    full_csv_url = urljoin(url, csv_link)
    logging.info(f'Downloading the CSV file from {full_csv_url}...')
    
    csv_response = requests.get(full_csv_url)
    df = pd.read_csv(BytesIO(csv_response.content))
    df.columns = df.columns.str.strip()

    pattern = re.compile(r'BRENT CRUDE|WTI CRUDE', re.IGNORECASE)
    filtered_df = df[df['constituent_description'].str.contains(pattern, na=False)][['constituent_description', 'shares_held_of_constituent', 'constituent_weight']]
    filtered_df.columns = ['name', 'quantity', 'weight']
    filtered_df['date'] = datetime.now(local_tz).strftime('%Y-%m-%d')
    filtered_df['ETF'] = 'American Beacon'
    return filtered_df
    
def scrape_imgp_funds():
    logging.info('Starting the scraping process for IMGP Funds...')
    url = 'https://imgpfunds.com/im-dbi-managed-futures-strategy-etf/'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    link_tag = soup.find('a', href=re.compile(r'DBMF.*?\.xlsx'))
    if not link_tag:
        logging.warning("No DBMF Holdings Excel file found on the IMGP Funds site.")
        return None  
    excel_link = link_tag['href']

    full_excel_url = urljoin(url, excel_link)
    logging.info(f'Downloading the Excel file from {full_excel_url}...')

    excel_response = requests.get(full_excel_url)
    df = pd.read_excel(BytesIO(excel_response.content), skiprows=5)
    df.columns = df.columns.str.strip()

    pattern = re.compile(r'BRENT CRUDE|WTI CRUDE', re.IGNORECASE)
    filtered_df = df[df['DESCRIPTION'].str.contains(pattern, na=False)][['DESCRIPTION', 'SHARES', 'PCT_HOLDINGS']]
    filtered_df.columns = ['name', 'quantity', 'weight']
    filtered_df['date'] = datetime.now(local_tz).strftime('%Y-%m-%d')
    filtered_df['ETF'] = 'IMGP Funds'
    return filtered_df

def scrape_kfa_funds():
    logging.info('Starting the scraping process for KFA Funds...')
    url = 'https://kfafunds.com/kmlm/#holdings'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }
    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')
    csv_link_tag = soup.find('a', id='full_holdings')
    csv_link = csv_link_tag['href']
    logging.info(f'Found CSV link: {csv_link}')
    csv_response = requests.get(csv_link, headers=headers)
    df = pd.read_csv(BytesIO(csv_response.content), skiprows=1)
    df.columns = df.columns.str.strip()

    pattern = re.compile(r'GASOLINE RBOB|WTI CRUDE FUTURE|NY HARB ULSD', re.IGNORECASE)
    filtered_df = df[df['Company Name'].str.contains(pattern, na=False)][['Company Name', 'Shares Held', '% of Net Assets']]
    filtered_df.columns = ['name', 'quantity', 'weight']
    filtered_df['date'] = datetime.now(local_tz).strftime('%Y-%m-%d')
    filtered_df['ETF'] = 'KFA Funds'
    return filtered_df

def scrape_simplify():
    logging.info("Starting the scraping process for Simplify...")
    url = 'https://www.simplify.us/etfs/cta-simplify-managed-futures-strategy-etf#portfolio-holdings'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    rows = soup.find_all('tr')

    patterns = [r'wti crude future', r'ny harb ulsd', r'gasoline']
    combined_pattern = re.compile('|'.join(patterns), re.IGNORECASE)

    data = []
    for row in rows:
        cells = row.find_all('td')
        if len(cells) >= 4:
            ticker = cells[0].get_text(strip=True).replace('Ticker', '').strip()
            name = cells[1].get_text(strip=True).replace('Name', '').strip()
            quantity = cells[2].get_text(strip=True).replace('Quantity', '').strip()
            weight = cells[3].get_text(strip=True).replace('Weight', '').strip()

            if combined_pattern.search(name):
                try:
                    quantity = int(quantity.replace(',', ''))
                except ValueError:
                    logging.warning(f"Could not convert quantity to int: {quantity}")
                    continue

                data.append({
                    'ticker': ticker,
                    'name': name,
                    'quantity': quantity,
                    'weight': weight,
                    'date': datetime.now(local_tz).strftime('%Y-%m-%d'),
                    'ETF': 'Simplify'
                })

    return pd.DataFrame(data)

def scrape_all_data():
    data_frames = [
        scrape_american_beacon(),
        scrape_imgp_funds(),
        scrape_kfa_funds(),
        scrape_simplify()
    ]
    combined_df = pd.concat(data_frames, ignore_index=True)
    return combined_df

def download_excel_from_s3():
    try:
        logging.info(f"Downloading {file_key} from S3 bucket {bucket_name}...")
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        return BytesIO(response['Body'].read())
    except Exception as e:
        logging.error(f"An error occurred while downloading the Excel file from S3: {str(e)}")
        logging.error(traceback.format_exc())
        return None

def upload_excel_to_s3(file_content):
    try:
        logging.info(f"Uploading {file_key} to S3 bucket {bucket_name}...")
        s3_client.put_object(Bucket=bucket_name, Key=file_key, Body=file_content)
        logging.info("Upload successful.")
    except Exception as e:
        logging.error(f"An error occurred while uploading the Excel file to S3: {str(e)}")
        logging.error(traceback.format_exc())

def update_excel(data):
    changes = []
    logging.info(f"Updating Excel file: {file_key}")
    try:
        excel_file = download_excel_from_s3()
        if excel_file:
            workbook = load_workbook(excel_file)
        else:
            workbook = Workbook()
            workbook.remove(workbook.active)  # Remove default sheet

        for _, entry in data.iterrows():
            entry['ETF'] = entry['ETF'].strip()  # Clean ETF name
            entry['name'] = entry['name'].strip()  # Clean name

            # Create a simplified sheet name
            etf_abbreviation = ETF_ABBREVIATIONS.get(entry['ETF'], entry['ETF'])
            contract_type = ''
            expiration = ''

            if 'WTI CRUDE' in entry['name'].upper():
                contract_type = 'WTI'
            elif 'BRENT CRUDE' in entry['name'].upper():
                contract_type = 'BRENT'
            elif 'GASOLINE' in entry['name'].upper() or 'RBOB' in entry['name'].upper():
                contract_type = 'RBOB'
            elif 'ULSD' in entry['name'].upper() or 'HEATING OIL' in entry['name'].upper():
                contract_type = 'HO'

            # Try to extract expiration date (assuming it's in the format MMMYY)
            date_match = re.search(r'(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\d{2}', entry['name'].upper())
            if date_match:
                expiration = date_match.group()

            sheet_name = f"{etf_abbreviation} {contract_type} {expiration}".strip()
            sheet_name = sheet_name[:31]  # Excel has a 31 character limit for sheet names

            if sheet_name not in workbook.sheetnames:
                logging.info(f"Creating new sheet: {sheet_name}")
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(['Date', 'ETF', 'Name', 'Quantity', 'Weight'])
            else:
                sheet = workbook[sheet_name]

            # Check for quantity changes
            if sheet.max_row > 1:
                prev_quantity = sheet.cell(row=sheet.max_row, column=4).value
                if prev_quantity is not None:
                    prev_quantity = int(prev_quantity)
                    current_quantity = int(entry['quantity'])
                    if prev_quantity != current_quantity:
                        changes.append((entry['ETF'], entry['name'], prev_quantity, current_quantity))

            # Append the data
            entry_data = [entry['date'], entry['ETF'], entry['name'], entry['quantity'], entry['weight']]
            sheet.append(entry_data)

        # Save the workbook to a BytesIO object
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        upload_excel_to_s3(excel_buffer)

        return changes, excel_buffer
    except Exception as e:
        logging.error(f"An error occurred while updating Excel: {str(e)}")
        logging.error(traceback.format_exc())
        return [], None

def create_visualizations(excel_buffer):
    logging.info("Creating visualizations...")
    plot_buffers = []
    try:
        df = pd.read_excel(excel_buffer, sheet_name=None)
        color_scheme = {
            'WTI': 'blue',
            'RBOB': 'green',
            'HO': 'red',
            'BRENT': 'orange'
        }
        for etf in ETF_ABBREVIATIONS.values():
            etf_data = {}
            for sheet_name, sheet_df in df.items():
                if sheet_name.startswith(etf):
                    contract_type = sheet_name.split()[1]
                    if contract_type not in etf_data:
                        etf_data[contract_type] = []
                    etf_data[contract_type].extend(sheet_df[['Date', 'Quantity']].values.tolist())
            if etf_data:
                plt.figure(figsize=(12, 6))
                for contract_type, data in etf_data.items():
                    data_df = pd.DataFrame(data, columns=['Date', 'Quantity'])
                    data_df['Date'] = pd.to_datetime(data_df['Date'])
                    data_df = data_df.groupby('Date')['Quantity'].sum().reset_index()
                    data_df = data_df.sort_values('Date')
                    color = color_scheme.get(contract_type, 'black')
                    plt.plot(data_df['Date'], data_df['Quantity'], label=contract_type, marker='o', color=color)
                
                plt.title(f'Contract Positions for {etf}')
                plt.xlabel('Date')
                plt.ylabel('Quantity')
                plt.legend()
                plt.xticks(rotation=45)
                plt.grid(True, linestyle='--', alpha=0.7)
                
                if etf == 'DBMF':
                    ax = plt.gca()
                    ax.yaxis.set_major_formatter(ScalarFormatter(useOffset=False))
                    ax.ticklabel_format(style='plain', axis='y')
                
                plt.tight_layout()
                img_buffer = BytesIO()
                plt.savefig(img_buffer, format='png')
                img_buffer.seek(0)
                plot_buffers.append((f'{etf}_positions', img_buffer))
                plt.close()
        logging.info("Visualizations created successfully.")
        return plot_buffers
    except Exception as e:
        logging.error(f"An error occurred while creating visualizations: {str(e)}")
        logging.error(traceback.format_exc())
        return []

def clean_name(name):
    if isinstance(name, str):
        match = re.search(r"(WTI|BRENT|RBOB|HO)", name)
        return match.group(0) if match else None
    return None

def calculate_crack_spread(df):
    for col in ['WTI', 'BRENT', 'RBOB', 'HO']:
        if col not in df.columns:
            df[col] = 0
    
    df['Adjusted Quantity'] = (
        (1 * (df['WTI'])) + 
        (1 * (df['BRENT'])) +
        (1 * df['RBOB']) +
        (1 * df['HO'])
    )
    return df

def custom_scale(df):
    max_abs_value = df['Adjusted Quantity'].abs().max()
    df['Scaled Quantity'] = df['Adjusted Quantity'] / max_abs_value
    return df

def load_and_process_data(excel_buffer):
    xls = pd.ExcelFile(excel_buffer)
    etf_data = {'AHL': [], 'DBMF': [], 'KMLM': [], 'SY': []}
    
    for sheet_name in xls.sheet_names:
        etf_name = sheet_name.split()[0]
        
        if etf_name in etf_data:
            df = pd.read_excel(xls, sheet_name)
            df['Date'] = pd.to_datetime(df['Date'])
            df['Commodity'] = df['Name'].apply(clean_name)
            df['Quantity'] = df['Quantity'].astype(float)

            df = df[df['Commodity'].notnull()]
            df = df[['Date', 'Commodity', 'Quantity']]
            etf_data[etf_name].append(df)
    
    for etf in etf_data:
        if etf_data[etf]:
            df = pd.concat(etf_data[etf])
            df_pivot = df.pivot_table(index='Date', columns='Commodity', values='Quantity', aggfunc='sum').reset_index()
            df_pivot.fillna(0, inplace=True)
            
            df_pivot = calculate_crack_spread(df_pivot)
            df_pivot = custom_scale(df_pivot)
            
            if etf in ['AHL', 'DBMF']:
                df_pivot['Adjusted Quantity'] /= 1000
            
            etf_data[etf] = df_pivot[['Date', 'Adjusted Quantity', 'Scaled Quantity']]
    
    return etf_data

def create_weighted_proxy(etf_data, weights, use_scaled=False):
    proxy = pd.DataFrame()
    quantity_col = 'Scaled Quantity' if use_scaled else 'Adjusted Quantity'
    
    for etf, weight in weights.items():
        if proxy.empty:
            proxy = etf_data[etf].copy()
            proxy[f'Weighted_{etf}'] = proxy[quantity_col] * weight
        else:
            proxy = proxy.merge(etf_data[etf], on='Date', suffixes=('', f'_{etf}'))
            proxy[f'Weighted_{etf}'] = proxy[f'{quantity_col}_{etf}'] * weight
    
    proxy['Total'] = proxy[[col for col in proxy.columns if col.startswith('Weighted_')]].sum(axis=1)
    return proxy

def plot_results(proxy, title, ylabel, only_total_line=False):
    plt.figure(figsize=(12, 6))
    if not only_total_line:
        for col in proxy.columns:
            if col.startswith('Weighted_'):
                plt.plot(proxy['Date'], proxy[col], label=col)
    plt.plot(proxy['Date'], proxy['Total'], label='Total Weighted Proxy', linewidth=2, color='black')
    plt.legend()
    plt.title(title)
    plt.xlabel('Date')
    plt.ylabel(ylabel)
    plt.tight_layout()
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png')
    img_buffer.seek(0)
    plt.close()
    return img_buffer


def main():
    data = scrape_all_data()
    changes, excel_buffer = update_excel(data)
    
    if excel_buffer:
        existing_plots = create_visualizations(excel_buffer)
        
        etf_data = load_and_process_data(excel_buffer)
        weights = {'AHL': 0.05, 'DBMF': 0.5, 'KMLM': 0.05, 'SY': 0.4}
        
        proxy1 = create_weighted_proxy(etf_data, weights, use_scaled=False)
        proxy2 = create_weighted_proxy(etf_data, weights, use_scaled=True)
        
        plot1 = plot_results(proxy1, 'CTA Proxy: Weighted ETFs with Division Adjustment', 'Weighted Adjusted Quantity')
        # For the second plot, set only_total_line=True to only plot the black line.
        plot2 = plot_results(proxy2, 'CTA Proxy: Weighted ETFs with Custom Scaling and 1:1 Ratio', 'Scaled Weighted Value', only_total_line=True)
        
        all_plots = existing_plots + [('proxy1', plot1), ('proxy2', plot2)]
        
        send_email(excel_buffer, changes, all_plots)
    else:
        logging.error("Failed to update Excel file. Email not sent.")

def send_email(excel_buffer, changes, all_plots):
    logging.info("Preparing to send email...")
    sender_email = "dsierraramirez115@gmail.com"
    receiver_email = ["diegosierra01@yahoo.com", "arnav.ashruchi@gmail.com","jordan.valer@lmrpartners.com"]
    password = os.environ['EMAIL_PASSWORD']

    message = MIMEMultipart("related")
    message["From"] = sender_email
    message["To"] = ", ".join(receiver_email)
    message["Subject"] = "Daily Scraped Data and CTA Proxies"

    html = """
    <html>
    <body>
    <p>Please find attached the daily scraped data and visualizations.</p>
    """

    if changes:
        html += "<p>Changes in quantities recorded:</p><ul>"
        for change in changes:
            html += f"<li>{change[0]} - {change[1]}: {change[2]} to {change[3]}</li>"
        html += "</ul>"
    else:
        html += "<p>No changes in quantities recorded.</p>"

    html += "<h2>Visualizations</h2>"
    
    for plot_name, _ in all_plots:
        html += f'<img src="cid:{plot_name}"><br><br>'

    html += "</body></html>"

    message.attach(MIMEText(html, "html"))

    for plot_name, plot_buffer in all_plots:
        img = MIMEImage(plot_buffer.getvalue())
        img.add_header('Content-ID', f'<{plot_name}>')
        message.attach(img)

    excel_attachment = MIMEApplication(excel_buffer.getvalue(), _subtype="xlsx")
    excel_attachment.add_header('Content-Disposition', 'attachment', filename="4scrapesdata.xlsx")
    message.attach(excel_attachment)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, password)
            server.send_message(message)
        logging.info("Email sent successfully.")
    except Exception as e:
        logging.error(f"An error occurred while sending the email: {str(e)}")
        logging.error(traceback.format_exc())

if __name__ == "__main__":
    main()
